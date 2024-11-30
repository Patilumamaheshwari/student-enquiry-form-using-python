from tkinter import *
from  tkinter import messagebox
from openpyxl import Workbook
import pathlib
import openpyxl,xlrd
from datetime import datetime
import pymysql  



###################################################### create database ###############################################################

connection = pymysql.connect(host="localhost",user="root",passwd="",database="Student")
cursor = connection.cursor()
# print("connected"/)


########################################################## SAVE ######################################################################################
def save():
    status=messagebox.askyesno(title="question",message="Do u want save ")
    if status==True:
        N=name.get()
        d=date.get()
        e=email.get()
        p=phonenumber.get()
        ad=address.get()
        a=alternatephonenumber.get()
        co=course.get()
        b=batch.get()
        h=come.get()
        fe=fresher.get()
        pe=person.get()
        cu=counselor.get()
        fa=fee.get()
        cm=comment.get()
        if N==""or d==""or e==""or p=="" or ad=="" or a=="" or co=="" or b=="" or h=="" or fe=="" or pe=="" or cu=="" or fa=="" or cm=="":
            messagebox.showerror("error","few data is missing")
        else:
            file=openpyxl.load_workbook('student_data.xlsx')
            sheet=file.active
            sheet.cell(column=1,row=sheet.max_row+1,value=N)
            sheet.cell(column=2,row=sheet.max_row,value=d)
            sheet.cell(column=3,row=sheet.max_row,value=e)
            sheet.cell(column=4,row=sheet.max_row,value=p)
            sheet.cell(column=5,row=sheet.max_row,value=ad)
            sheet.cell(column=6,row=sheet.max_row,value=a)
            sheet.cell(column=7,row=sheet.max_row,value=co)
            sheet.cell(column=8,row=sheet.max_row,value=b)
            sheet.cell(column=9,row=sheet.max_row,value=h)
            sheet.cell(column=10,row=sheet.max_row,value=fe)
            sheet.cell(column=11,row=sheet.max_row,value=pe)
            sheet.cell(column=12,row=sheet.max_row,value=cu)
            sheet.cell(column=13,row=sheet.max_row,value=fa)
            sheet.cell(column=14,row=sheet.max_row,value=cm)


            if reg.get() == 1 and enq.get() == 1:
                insert_reg_record = "INSERT INTO Registration(Date,Name,Mobile_Number,Alternate_Number,Email_Id,Address,Course_Intersted,Batch_Preferred,How_You_Came_To_Know_us,Experience_Fresher,Contact_Person_From_Besant,Counselor,Fees,Comments) VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s');"%(date.get(),name.get(),phonenumber.get(),alternatephonenumber.get(),email.get(),address.get(),course.get(),batch.get(),come.get(),fresher.get(),person.get(),counselor.get(),fee.get(),comment.get())
                insert_enq_record = "INSERT INTO Enquiry(Date,Name,Mobile_Number,Alternate_Number,Email_Id,Address,Course_Intersted,Batch_Preferred,How_You_Came_To_Know_us,Experience_Fresher,Contact_Person_From_Besant,Counselor,Fees,Comments) VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s');"%(date.get(),name.get(),phonenumber.get(),alternatephonenumber.get(),email.get(),address.get(),course.get(),batch.get(),come.get(),fresher.get(),person.get(),counselor.get(),fee.get(),comment.get())
            elif reg.get() == 1:
                insert_reg_record = "INSERT INTO Registration(Date,Name,Mobile_Number,Alternate_Number,Email_Id,Address,Course_Intersted,Batch_Preferred,How_You_Came_To_Know_us,Experience_Fresher,Contact_Person_From_Besant,Counselor,Fees,Comments) VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s');"%(date.get(),name.get(),phonenumber.get(),alternatephonenumber.get(),email.get(),address.get(),course.get(),batch.get(),come.get(),fresher.get(),person.get(),counselor.get(),fee.get(),comment.get())
            elif enq.get() == 1:
                insert_enq_record = "INSERT INTO Enquiry(Date,Name,Mobile_Number,Alternate_Number,Email_Id,Address,Course_Intersted,Batch_Preferred,How_You_Came_To_Know_us,Experience_Fresher,Contact_Person_From_Besant,Counselor,Fees,Comments) VALUES('%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s','%s');"%(date.get(),name.get(),phonenumber.get(),alternatephonenumber.get(),email.get(),address.get(),course.get(),batch.get(),come.get(),fresher.get(),person.get(),counselor.get(),fee.get(),comment.get())
                print("hello")
            else:
                print("data is missing")
            if reg.get() == 1:
                cursor.execute(insert_reg_record) 
            if  enq.get() == 1:
                cursor.execute(insert_enq_record) 
            if reg.get() == 1 or enq.get() == 1:
                connection.commit()
            file.save(r'student_data.xlsx')
            
        




        

####################################################################### DELETE ###################################################################
def delet():
    status=messagebox.askyesno(title="question",message="Do u want delete ")
    if status==True:
        window.destroy()
    else:
        pass




########################################################3 File creatiion #############################################
file=pathlib.Path("student_data.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Name"
    sheet['B1']="Date"
    sheet['C1']="Email"
    sheet['D1']="Phone Number"
    sheet['E1']="Address"
    sheet['F1']="Alternate Phone Number"
    sheet['G1']="Course Intrested"
    sheet['H1']="Batch Prefered"
    sheet['I1']="How You Come To Know Us"
    sheet['J1']="Are You Exprience Or Fresher"
    sheet['K1']="Contact Person From Besant Technology"
    sheet['L1']="Counselor"
    sheet['M1']="Fee"
    sheet['N1']="Comment"

    file.save("student_data.xlsx")

    
    




################################################################ module ##################################################
window=Tk()
window.title("student form")
window.geometry("1000x1000")
window.configure(bg="lightblue")

title=Label(window,text="BESANT TECHNOLOGY",fg="red",bg="lightblue",font="bold")
title.grid(row=0,column=1)

name=Label(window,text="Name",font="bold",bg="lightblue")
name.grid(row=1,column=0,sticky=W)
date=Label(window,text="Date",font="bold",bg="lightblue")
date.grid(row=2,column=0,sticky=W)
email=Label(window,text="Email",font="bold",bg="lightblue")
email.grid(row=3,column=0,sticky=W)
phonenumber=Label(window,text="Phone Number",font="bold",bg="lightblue")
phonenumber.grid(row=4,column=0,sticky=W)
address=Label(window,text="Address",font="bold",bg="lightblue")
address.grid(row=6,column=0,sticky=W)
alternatephonenumber=Label(window,text=" Alternate Phone Number",font="bold",bg="lightblue")
alternatephonenumber.grid(row=7,column=0,sticky=W)
course=Label(window,text="Course Intrested",font="bold",bg="lightblue")
course.grid(row=8,column=0,sticky=W)
batch=Label(window,text="Batch Prefered",font="bold",bg="lightblue")
batch.grid(row=9,column=0,sticky=W)
come=Label(window,text="How You Come To Know Us",font="bold",bg="lightblue")
come.grid(row=10,column=0,sticky=W)
fresher=Label(window,text="Are You Exprience Or Fresher",font="bold",bg="lightblue")
fresher.grid(row=11,column=0,sticky=W)
person=Label(window,text="Contact Person From Besant Technology",font="bold",bg="lightblue")
person.grid(row=12,column=0,sticky=W)
counselor=Label(window,text="Counselor",font="bold",bg="lightblue")
counselor.grid(row=13,column=0,sticky=W)
fee=Label(window,text="Fee",font="bold",bg="lightblue")
fee.grid(row=14,column=0,sticky=W)
comment=Label(window,text="Comment",font="bold",bg="lightblue")
comment.grid(row=15,column=0,sticky=W)

################################################################################### data entry ###############################################################################
name=Entry(window,width=40)
name.grid(row=1,column=1)
date=Entry(window,width=40)
date.grid(row=2,column=1)
email=Entry(window,width=40)
email.grid(row=3,column=1)
phonenumber=Entry(window,width=40)
phonenumber.grid(row=4,column=1)
address=Entry(window,width=40)
address.grid(row=6,column=1)
alternatephonenumber=Entry(window,width=40)
alternatephonenumber.grid(row=7,column=1)
course=Entry(window,width=40)
course.grid(row=8,column=1)
batch=Entry(window,width=40)
batch.grid(row=9,column=1)
come=Entry(window,width=40)
come.grid(row=10,column=1)
fresher=Entry(window,width=40)
fresher.grid(row=11,column=1)
person=Entry(window,width=40)
person.grid(row=12,column=1)
counselor=Entry(window,width=40)
counselor.grid(row=13,column=1)
fee=Entry(window,width=40)
fee.grid(row=14,column=1)
comment=Entry(window,width=40)
comment.grid(row=15,column=1)
################################################# buttons #######################################################
reg=IntVar()
enq=IntVar()
ceck=Checkbutton(window,text="Registration",variable=reg).grid(row=16,column=1,sticky=W,padx="10",pady="10")
ceck=Checkbutton(window,text="Enquiry",variable=enq).grid(row=16,column=1,sticky=E)
savee=Button(window,text="Submit",command=save,font="bold",bg="Green")
savee.grid(row=20,column=1,sticky=W)
savee=Button(window,text="Delete",command=delet,font="bold",bg="red")
savee.grid(row=20,column=2,sticky=N)

window.mainloop()
